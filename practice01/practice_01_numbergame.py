'''
设计一个猜数字的程序，在0-100的整数范围内，随机生成一个整数，输入数字判断是否和生成的数据一致。
返回猜测结果是是大了还是小了输出，直到猜中，程序结束。
需要将之前测的数据做记录最终以列表的形式输出。
设计思路：
1、程序分两部分，第一部分，是对比数字大小；第二部分是结果写入excel
2、对比数字大小：
step1:用input获取输入，得到的都是字符串，写入第row行第一列
step2:强转输入的字符串，捕获异常，强转成功，则进行step3，否则重新进行step2 row = row+1
step3:输入的数字和随机生成的数字进行比较，有4种情况，
a:小于0或者大于100，提示不在随机数的范围内，猜数游戏范围为[0,100] ，猜题结果写入第row行第二列，跳到step2 row = row+1；
b：比随机数大，提示猜大了，猜题结果写入第row行第二列，跳到step2 row = row + 1；
c:比随机数小，提示猜小了，猜题结果写入第row行第二列，跳到step2 row = row + 1；
d：和随机数相等，猜对答案，猜题结果写入第row行第二列 游戏结束
step4:保存excel

'''
import logging
import random
from openpyxl import Workbook
import sys
import io
import unittest
from prettytable import *
logging.basicConfig(level=logging.INFO)
class Game():
    def __init__(self):
        pass

    def guessnumber(self):
        answer = random.randint(0,100)
        x = PrettyTable()
        x.field_names = ["times", "Your Guess", "Result"]
        print("该回合数字迷底为：%s" % answer)
        wb = Workbook()
        ws = wb.active
        row = 1
        while answer:
            try:
                num = input("请输入：")
                ws.cell(row,1).value = num
                guess = int(num)
                if guess == answer:
                    print("答对了")
                    ws.cell(row,2).value = "回答正确" #写到excel里
                    x.add_row([row,num,"回答正确"]) #作为待会要输出的列表内容
                    break
                if guess > answer and guess <= 100:
                    print("回答错误，猜的数太大了")
                    ws.cell(row,2).value = "回答错误，猜大了"
                    x.add_row([row, num, "回答错误，猜大了"])
                if guess < answer and guess >= 0:
                    print("回答错误，猜的数太小了")
                    ws.cell(row,2).value = "回答错误，猜小了"
                    x.add_row([row, num, "回答错误，猜小了"])
                if guess < 0 or guess > 100:
                    print("超出猜数范围了哦")
                    ws.cell(row,2).value = "回答错误，超出范围了"
                    x.add_row([row, num, "回答错误，超出范围"])
                if row > 15:
                    print("你太笨了，猜题机会已经全部用完了")
                    break
                row = row + 1
            except ValueError as e:
                print("输入的不是整数")
                # ws.cell(row, 1).value = num
                ws.cell(row,2).value = "回答错误，输入的不是整数"
                x.add_row([row,num,"回答错误，输入的不是整数"])
                row = row + 1
        wb.save("猜题结果.xls")
        x.set_style(PLAIN_COLUMNS)  #控制台打印的输出列表不显示边框
        print(x)


if __name__ == '__main__':
    Game().guessnumber()




