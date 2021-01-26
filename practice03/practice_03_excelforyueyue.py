'''
step1:读excel,把表格读成嵌套字典{'张三'：{'早会':'√','追踪 (单)':10,'意向（单）':15,'报价（单） ':1,'意向（单）':15,...},'李四':{}}
step2：在需要写入的表格中检索对应的专员，然后把专员对应的一些需要填写的值填入表格，并保存

这个题目没有做完。。。。只填了部分，其他的值不知道干啥的，填写的对应关系也不是很懂
'''

from copy import deepcopy
import json
from openpyxl import load_workbook


class Read_Ex():
    def read_excel_openpyxl(self):
        wb = load_workbook('dddd.xlsx')
        wb_sheets = wb.sheetnames
        print(wb_sheets)
        s = {}
        print(wb[wb_sheets[1]].max_row)
        print(wb[wb_sheets[1]].max_column)
        key1 = [] #第一行数据做为key值
        key2 = [] #第一列数据
        for col in wb[wb_sheets[1]].iter_rows(min_row=2,min_col=3,max_col=14,max_row=2):
            for cell in col:
                key1.append(cell.value)
        col_Num = len(key1) + 1 #最大列数
        for row in wb[wb_sheets[1]].iter_cols(min_col=2,min_row=3,max_row=12,max_col=2):
            for cell in row:
                key2.append(cell.value)
        row_Num = len(key2) + 1 #最大行数
        print('key1是：', key1)
        print('key2是：',key2)
        if row_Num <= 1:
            print('没数据')
        else:
            y = 0
            i = 3
            j = 0
            while i <= row_Num and y <= row_Num:
                for x in wb[wb_sheets[1]].iter_rows(min_row=i,min_col=3,max_col=14,max_row=i):
                    # print('x是以下这些单元格：',x)
                    d = {}
                    # print(len(x))
                    for j in range(0,col_Num-1):
                        d[key1[j]] = x[j].value
                    s[key2[y]] = d
                    y = y + 1
                    i = i + 1
            print(json.dumps(s,indent=3,ensure_ascii=False))
            return s


    def write_excel(self):
        # 打开excel表，填写路径
        data = self.read_excel_openpyxl()
        wb = load_workbook("dddd.xlsx")
        all_sheets = wb.sheetnames
        # 找到sheet页
        ws = wb[all_sheets[0]]
        for i in data.keys():
            row,col = self.search_value(i)
            ws.cell(row,col+8).value = data[i]['早会']
            ws.cell(row,col+1).value = data[i]['追踪 (单）']
            ws.cell(row,col+2).value = data[i]['意向（单） ']
            ws.cell(row,col+3).value = data[i]['报价（单） ']
            ws.cell(row,col+4).value = data[i]['转保意向（单） ']
            wb.save('dddd1.xlsx')

    #获得取消合并单元格的列表
    def get_merge_cell(self):
        wb = load_workbook('dddd.xlsx')
        all_sheets = wb.sheetnames
        ws = wb[all_sheets[0]]
        #复制需要操作的工作表
        ws_bake = wb.copy_worksheet(ws)
        #获取表格中的合并单元格的位置
        m_list = ws_bake.merged_cells
        cr = []

        for m_area in m_list:
            r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
            # 纵向合并单元格的位置信息提取出
            if r2 - r1 > 0:
                cr.append((r1, r2, c1, c2))
        print(cr)
            # 这里注意需要把合并单元格的信息提取出再拆分
        merge_cr = deepcopy(cr)
        for r in cr:
            ws.unmerge_cells(start_row=r[0], end_row=r[20], start_column=r[1], end_column=r[20])
        wb.save('dddd.xlsx')




    #在表格中找到对应值，并返回该单元格所在的行，列
    def search_value(self,keyword):
        wb = load_workbook('dddd.xlsx')
        all_sheets = wb.sheetnames
        sheet = wb[all_sheets[0]]
        # m_list = sheet.merged_cells
        # print(m_list)
        for row in sheet.iter_rows(min_row=2,min_col=2,max_col=15,max_row=17):
            for cell in row:
                t = cell.value
                if keyword == cell.value:
                    # print(cell.row)
                    # print(cell.column)
                    return cell.row,cell.column
                    # break


if __name__ == '__main__':
    s = Read_Ex()
    # s.search_value('八项基本动作 ')
    s.write_excel()

